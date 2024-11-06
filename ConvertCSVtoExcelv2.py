import pandas as pd
import csv
import os
from tqdm import tqdm
from openpyxl import load_workbook

# Yêu cầu người dùng nhập đường dẫn file
file_path = input("Please enter the path to the CSV file: ")

# Kiểm tra xem file có tồn tại hay không
if not os.path.exists(file_path):
    print("Error: The file was not found. Please check the path and try again.")
elif not file_path.lower().endswith('.csv'):
    print("Error: The file is not in CSV format.")
else:
    try:
        # Tự động xác định delimiter bằng Sniffer
        with open(file_path, mode='r', encoding='utf-8') as file:
            dialect = csv.Sniffer().sniff(file.read(1024))
            file.seek(0)  # Đặt lại con trỏ file về đầu

            # Yêu cầu người dùng nhập đường dẫn thư mục lưu
            output_dir = input("Please enter the directory to save the Excel file: ")

            # Kiểm tra xem thư mục có tồn tại không
            while not os.path.exists(output_dir) or not os.path.isdir(output_dir):
                print("Error: The directory does not exist. Please enter a valid directory.")
                output_dir = input("Please enter the directory to save the Excel file: ")

            # Yêu cầu người dùng nhập tên file
            output_file_name = input("Please enter the name for the Excel file (without extension): ")
            output_file = os.path.join(output_dir, f"{output_file_name}.xlsx")

            # Xác định kích thước chunk và số dòng tổng
            chunk_size = 10000  # Kích thước mỗi phần (chunk) là 10,000 dòng
            total_lines = sum(1 for _ in open(file_path, 'r', encoding='utf-8')) - 1  # Bỏ dòng tiêu đề
            num_chunks = (total_lines // chunk_size) + 1  # Số lượng chunk ước tính

            print("Converting CSV to Excel...")

            # Ghi từng chunk vào file Excel
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                with tqdm(total=100, desc="Processing", unit="%", unit_scale=True) as pbar:
                    for i, chunk in enumerate(pd.read_csv(file, delimiter=dialect.delimiter, chunksize=chunk_size)):
                        startrow = writer.sheets['Sheet1'].max_row if 'Sheet1' in writer.sheets else 0
                        chunk.to_excel(writer, index=False, header=writer.sheets.get('Sheet1') is None, startrow=startrow)
                        pbar.update((i + 1) / num_chunks * 100 - pbar.n)

            # Điều chỉnh độ rộng cột
            wb = load_workbook(output_file)
            ws = wb.active

            print("Adjusting column widths...")
            for col in tqdm(ws.columns, desc="Processing columns", unit="col"):
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[col_letter].width = adjusted_width

            # Lưu lại file sau khi điều chỉnh độ rộng cột
            wb.save(output_file)
            print("Conversion complete. File saved as", output_file)
    except Exception as e:
        print("An error occurred:", e)
