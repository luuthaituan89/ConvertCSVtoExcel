import pandas as pd
import csv
import os
from tqdm import tqdm
from openpyxl import load_workbook

# Yêu cầu người dùng nhập thư mục chứa các file CSV
folder_path = input("Please enter the directory containing the CSV files: ")

# Kiểm tra xem thư mục có tồn tại không
if not os.path.exists(folder_path) or not os.path.isdir(folder_path):
    print("Error: The directory was not found. Please check the path and try again.")
else:
    try:
        # Lấy danh sách file CSV trong thư mục
        csv_files = [f for f in os.listdir(folder_path) if f.lower().endswith('.csv')]

        if not csv_files:
            print("No CSV files found in the directory.")
        else:
            # Yêu cầu người dùng nhập thư mục lưu và tên file Excel
            output_dir = input("Please enter the directory to save the Excel file: ")
            while not os.path.exists(output_dir) or not os.path.isdir(output_dir):
                print("Error: The directory does not exist. Please enter a valid directory.")
                output_dir = input("Please enter the directory to save the Excel file: ")

            output_file_name = input("Please enter the name for the Excel file (without extension): ")
            output_file = os.path.join(output_dir, f"{output_file_name}.xlsx")

            print("Merging CSV files into a single Excel file...")

            # Ghi dữ liệu từ mỗi file CSV vào một sheet riêng trong Excel, với chunking
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                for csv_file in csv_files:
                    csv_path = os.path.join(folder_path, csv_file)
                    base_sheet_name = os.path.splitext(csv_file)[0][:28]  # Giới hạn tên sheet

                    # Xác định số dòng trong file CSV để tạo thanh tiến trình chính xác
                    total_rows = sum(1 for _ in open(csv_path, 'r', encoding='utf-8')) - 1  # Trừ dòng tiêu đề
                    chunk_size = 10000  # Kích thước mỗi chunk
                    num_chunks = (total_rows // chunk_size) + 1

                    # Tạo thanh tiến trình cho các chunk trong file
                    with tqdm(total=num_chunks, desc=f"Processing {csv_file}", unit="chunk") as pbar:
                        chunk_iter = pd.read_csv(csv_path, chunksize=chunk_size, delimiter=',', encoding='utf-8')
                        sheet_index = 1
                        current_sheet_name = f"{base_sheet_name}_Part{sheet_index}"
                        startrow = 0

                        for chunk in chunk_iter:
                            # Nếu số dòng hiện tại vượt quá giới hạn, chuyển sang sheet mới
                            if startrow + len(chunk) > 1048576:
                                sheet_index += 1
                                current_sheet_name = f"{base_sheet_name}_Part{sheet_index}"
                                startrow = 0  # Reset dòng bắt đầu cho sheet mới

                            # Ghi chunk vào sheet hiện tại
                            chunk.to_excel(writer, sheet_name=current_sheet_name, index=False, startrow=startrow,
                                           header=(startrow == 0))
                            startrow += len(chunk)  # Cập nhật dòng bắt đầu cho chunk tiếp theo
                            pbar.update(1)  # Cập nhật tiến trình cho mỗi chunk

            print("All CSV files merged successfully. Starting column width adjustment...")

            # Sau khi hoàn tất gộp dữ liệu, thực hiện điều chỉnh độ rộng cột
            wb = load_workbook(output_file)
            total_sheets = len(wb.sheetnames)

            # Sử dụng tqdm để hiển thị tiến độ khi điều chỉnh độ rộng cột
            with tqdm(total=total_sheets, desc="Adjusting column widths", unit="sheet") as pbar:
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    for col in ws.columns:
                        max_length = 0
                        col_letter = col[0].column_letter
                        for cell in col:
                            try:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                        adjusted_width = max_length + 2
                        ws.column_dimensions[col_letter].width = adjusted_width
                    pbar.update(1)  # Cập nhật tiến trình sau mỗi sheet

            # Lưu file sau khi điều chỉnh độ rộng cột
            wb.save(output_file)
            print(f"Merge and column adjustment complete. File saved as {output_file}")

    except Exception as e:
        print("An error occurred:", e)
